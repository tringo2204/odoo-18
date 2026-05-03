import base64
import io

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
except ImportError:
    openpyxl = None


class HrVnSiD02Report(models.Model):
    _name = 'hr.vn.si.d02.report'
    _description = 'Báo cáo D02-LT (Biến động lao động)'
    _inherit = ['mail.thread']
    _order = 'year desc, month desc'
    _check_company_auto = True

    name = fields.Char(
        string='Số báo cáo', readonly=True, copy=False, default='Mới',
    )
    month = fields.Selection([
        ('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'),
        ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
        ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'),
        ('10', 'Tháng 10'), ('11', 'Tháng 11'), ('12', 'Tháng 12'),
    ], string='Tháng', required=True, tracking=True)
    year = fields.Integer(string='Năm', required=True, tracking=True)
    company_id = fields.Many2one(
        'res.company', string='Công ty',
        default=lambda self: self.env.company,
    )
    monthly_list_id = fields.Many2one(
        'hr.vn.si.monthly.list', string='DS tăng/giảm tháng',
        readonly=True,
    )
    line_ids = fields.One2many(
        'hr.vn.si.d02.line', 'report_id', string='Chi tiết biến động',
    )
    line_count = fields.Integer(
        string='Số dòng', compute='_compute_line_count',
    )
    state = fields.Selection([
        ('draft', 'Nháp'),
        ('confirmed', 'Đã xác nhận'),
        ('submitted', 'Đã nộp'),
    ], string='Trạng thái', default='draft', tracking=True)
    export_file = fields.Binary(string='File xuất', readonly=True)
    export_filename = fields.Char(string='Tên file')
    note = fields.Text(string='Ghi chú')

    @api.depends('line_ids')
    def _compute_line_count(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'Mới') == 'Mới':
                vals['name'] = self.env['ir.sequence'].next_by_code(
                    'hr.vn.si.d02.report',
                ) or 'Mới'
        return super().create(vals_list)

    def action_confirm(self):
        for rec in self:
            if rec.state != 'draft':
                raise UserError(
                    _('Chỉ xác nhận báo cáo ở trạng thái Nháp.')
                )
            if not rec.line_ids:
                raise UserError(_('Báo cáo chưa có dòng chi tiết.'))
        self.write({'state': 'confirmed'})

    def action_submit(self):
        for rec in self:
            if rec.state != 'confirmed':
                raise UserError(
                    _('Phải xác nhận báo cáo trước khi nộp.')
                )
        self.write({'state': 'submitted'})

    def action_draft(self):
        for rec in self:
            if rec.state == 'submitted':
                raise UserError(
                    _('Không thể chuyển về Nháp khi đã nộp.')
                )
        self.write({'state': 'draft'})

    def action_generate_lines(self):
        """Tạo/cập nhật dòng D02 từ danh sách tăng/giảm liên kết."""
        self.ensure_one()
        if self.state != 'draft':
            raise UserError(_('Chỉ tạo dòng khi báo cáo ở trạng thái Nháp.'))
        if not self.monthly_list_id:
            raise UserError(_('Chưa liên kết danh sách tăng/giảm tháng.'))
        # Xóa dòng cũ
        self.line_ids.unlink()
        # Tạo dòng mới từ history confirmed
        history_entries = self.monthly_list_id.history_ids.filtered(
            lambda h: h.state in ('confirmed', 'reported')
            and h.change_type in ('increase', 'decrease', 'adjust', 'sick', 'maternity')
        )
        D02Line = self.env['hr.vn.si.d02.line']
        line_vals_list = []
        for entry in history_entries:
            line_vals_list.append({
                'report_id': self.id,
                'employee_id': entry.employee_id.id,
                'bhxh_number': entry.bhxh_number or '',
                'full_name': entry.employee_id.name,
                'change_type': entry.change_type,
                'old_salary': entry.old_salary,
                'new_salary': entry.new_salary,
                'effective_date': entry.effective_date,
            })
        lines = D02Line.create(line_vals_list)
        for entry, line in zip(history_entries, lines):
            entry._mark_reported(line)
        # Cập nhật trạng thái monthly list
        if self.monthly_list_id.state == 'confirmed':
            self.monthly_list_id.write({'state': 'exported'})

    def action_export_excel(self):
        """Xuất file D02 Excel và tải trực tiếp về máy."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('Vui lòng cài thư viện openpyxl: pip install openpyxl'))
        if not self.line_ids:
            raise UserError(_('Báo cáo chưa có dòng chi tiết. Hãy tạo dòng chi tiết trước.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'D02-LT'

        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        ws['A1'] = f'BIẾN ĐỘNG LAO ĐỘNG THAM GIA BHXH, BHYT, BHTN - THÁNG {self.month}/{self.year}'
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        headers = ['STT', 'Họ và tên', 'Mã số BHXH', 'Loại biến động',
                   'Mức lương cũ', 'Mức lương mới', 'Ngày hiệu lực', 'Ghi chú']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        change_labels = {
            'increase': 'Tăng', 'decrease': 'Giảm', 'adjust': 'Điều chỉnh',
            'sick': 'Ốm đau', 'maternity': 'Thai sản',
        }
        for idx, line in enumerate(self.line_ids, 1):
            r = idx + 2
            row_vals = [
                idx, line.full_name or line.employee_id.name,
                line.bhxh_number or '',
                change_labels.get(line.change_type, line.change_type),
                line.old_salary, line.new_salary,
                str(line.effective_date) if line.effective_date else '',
                '',
            ]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(r, col, val)
                c.border = thin
                if col in (5, 6):
                    c.number_format = '#,##0'

        col_widths = [6, 30, 15, 18, 18, 18, 15, 15]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        filename = f'D02-LT_T{self.month}_{self.year}.xlsx'
        file_data = base64.b64encode(output.getvalue())

        self.write({'export_file': file_data, 'export_filename': filename})

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self._name}/{self.id}/export_file/{filename}?download=true',
            'target': 'self',
        }
