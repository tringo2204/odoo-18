import base64
import io

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None


class SiBhxhExport(models.TransientModel):
    _name = 'hr.vn.si.bhxh.export'
    _description = 'Xuất mẫu BHXH cho cổng BHXH'

    report_type = fields.Selection([
        ('tk1_ts', 'TK1-TS — Tờ khai tham gia BHXH'),
        ('d02_ts', 'D02-TS — DS lao động tham gia BHXH'),
        ('d03_ts', 'D03-TS — DS đóng BHXH'),
    ], string='Loại báo cáo', required=True, default='d02_ts')
    month = fields.Selection([
        ('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'),
        ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
        ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'),
        ('10', 'Tháng 10'), ('11', 'Tháng 11'), ('12', 'Tháng 12'),
    ], string='Tháng', required=True)
    year = fields.Integer(
        string='Năm', required=True,
        default=lambda self: fields.Date.today().year,
    )
    company_id = fields.Many2one(
        'res.company', string='Công ty',
        default=lambda self: self.env.company,
    )
    export_file = fields.Binary(string='File xuất', readonly=True)
    export_filename = fields.Char()

    def action_export(self):
        self.ensure_one()
        method = {
            'tk1_ts': self._export_tk1_ts,
            'd02_ts': self._export_d02_ts,
            'd03_ts': self._export_d03_ts,
        }
        file_data, filename = method[self.report_type]()
        self.write({'export_file': file_data, 'export_filename': filename})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _get_active_records(self):
        """Lấy hồ sơ BH đang tham gia tại thời điểm tháng/năm đang xem."""
        month = int(self.month)
        year = self.year
        if month == 12:
            end_of_month = fields.Date.from_string(f'{year + 1}-01-01')
        else:
            end_of_month = fields.Date.from_string(f'{year}-{month + 1:02d}-01')
        return self.env['hr.vn.si.record'].search([
            ('current_status', 'in', ('active', 'suspended')),
            ('company_id', '=', self.company_id.id),
            ('registration_date', '<', end_of_month),
        ])

    def _export_tk1_ts(self):
        """TK1-TS: Tờ khai đơn vị tham gia BHXH."""
        if not openpyxl:
            raise UserError(_('Cần cài openpyxl.'))
        records = self._get_active_records()
        config = self.env['hr.vn.insurance.config'].search(
            [('year', '=', self.year)], limit=1,
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'TK1-TS'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:G1')
        ws['A1'] = f'TỜ KHAI ĐƠN VỊ THAM GIA BHXH, BHYT, BHTN — T{self.month}/{self.year}'
        ws['A1'].font = Font(bold=True, size=12)

        ws['A3'] = 'Đơn vị:'
        ws['B3'] = self.company_id.name
        ws['A4'] = 'Tổng số LĐ tham gia:'
        ws['B4'] = len(records)
        ws['A5'] = 'Tổng quỹ lương đóng:'
        ws['B5'] = sum(records.mapped('insurance_salary'))

        if config:
            ws['A7'] = 'Tỷ lệ đóng NV:'
            ws['B7'] = f'{config.bhxh_employee_rate + config.bhyt_employee_rate + config.bhtn_employee_rate}%'
            ws['A8'] = 'Tỷ lệ đóng DN:'
            ws['B8'] = f'{config.bhxh_employer_rate + config.bhyt_employer_rate + config.bhtn_employer_rate}%'

        output = io.BytesIO()
        wb.save(output)
        return base64.b64encode(output.getvalue()), f'TK1_TS_T{self.month}_{self.year}.xlsx'

    def _export_d02_ts(self):
        """D02-TS: Danh sách lao động tham gia BHXH."""
        if not openpyxl:
            raise UserError(_('Cần cài openpyxl.'))
        records = self._get_active_records()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'D02-TS'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:H1')
        ws['A1'] = f'DANH SÁCH LAO ĐỘNG THAM GIA BHXH — T{self.month}/{self.year}'
        ws['A1'].font = Font(bold=True, size=12)

        headers = ['STT', 'Họ và tên', 'Số sổ BHXH', 'Số thẻ BHYT',
                   'Phòng ban', 'Mức lương đóng', 'Nơi KCB', 'Ghi chú']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.border = thin

        for idx, rec in enumerate(records, 1):
            row = idx + 3
            ws.cell(row, 1, idx).border = thin
            ws.cell(row, 2, rec.employee_id.name).border = thin
            ws.cell(row, 3, rec.bhxh_number or '').border = thin
            ws.cell(row, 4, rec.bhyt_card_number or '').border = thin
            ws.cell(row, 5, rec.department_id.name or '').border = thin
            c = ws.cell(row, 6, rec.insurance_salary)
            c.border = thin
            c.number_format = '#,##0'
            ws.cell(row, 7, rec.bhyt_hospital_id or '').border = thin
            ws.cell(row, 8, '').border = thin

        for col in 'ABCDEFGH':
            ws.column_dimensions[col].width = 20

        output = io.BytesIO()
        wb.save(output)
        return base64.b64encode(output.getvalue()), f'D02_TS_T{self.month}_{self.year}.xlsx'

    def _export_d03_ts(self):
        """D03-TS: Danh sách đóng BHXH."""
        if not openpyxl:
            raise UserError(_('Cần cài openpyxl.'))
        records = self._get_active_records()
        config = self.env['hr.vn.insurance.config'].search(
            [('year', '=', self.year)], limit=1,
        )
        ee_rate = (config.bhxh_employee_rate + config.bhyt_employee_rate + config.bhtn_employee_rate) if config else 10.5
        er_rate = (config.bhxh_employer_rate + config.bhyt_employer_rate + config.bhtn_employer_rate) if config else 21.5

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'D03-TS'
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:I1')
        ws['A1'] = f'DANH SÁCH ĐÓNG BHXH, BHYT, BHTN — T{self.month}/{self.year}'
        ws['A1'].font = Font(bold=True, size=12)

        headers = ['STT', 'Họ và tên', 'Số sổ BHXH', 'Mức lương đóng',
                   'Phần NV đóng', 'Phần DN đóng', 'Tổng đóng',
                   'Phòng ban', 'Ghi chú']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.border = thin

        for idx, rec in enumerate(records, 1):
            row = idx + 3
            sal = rec.insurance_salary
            ee_amount = sal * ee_rate / 100
            er_amount = sal * er_rate / 100
            ws.cell(row, 1, idx).border = thin
            ws.cell(row, 2, rec.employee_id.name).border = thin
            ws.cell(row, 3, rec.bhxh_number or '').border = thin
            for col, val in enumerate([sal, ee_amount, er_amount, ee_amount + er_amount], 4):
                c = ws.cell(row, col, val)
                c.border = thin
                c.number_format = '#,##0'
            ws.cell(row, 8, rec.department_id.name or '').border = thin
            ws.cell(row, 9, '').border = thin

        for col in 'ABCDEFGHI':
            ws.column_dimensions[col].width = 18

        output = io.BytesIO()
        wb.save(output)
        return base64.b64encode(output.getvalue()), f'D03_TS_T{self.month}_{self.year}.xlsx'
