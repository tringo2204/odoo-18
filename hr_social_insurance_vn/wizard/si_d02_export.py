import base64
import io

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None


class SiD02Export(models.TransientModel):
    _name = 'hr.vn.si.d02.export'
    _description = 'Xuất báo cáo D02-LT'

    monthly_list_id = fields.Many2one(
        'hr.vn.si.monthly.list', string='DS tăng/giảm tháng',
        required=True,
    )
    month = fields.Selection([
        ('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'),
        ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
        ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'),
        ('10', 'Tháng 10'), ('11', 'Tháng 11'), ('12', 'Tháng 12'),
    ], string='Tháng', required=True)
    year = fields.Integer(string='Năm', required=True)

    def action_export(self):
        self.ensure_one()
        monthly_list = self.monthly_list_id
        if monthly_list.state != 'confirmed':
            raise UserError(
                _('Danh sách tăng/giảm phải ở trạng thái Đã xác nhận.')
            )

        # Tạo báo cáo D02
        report = self.env['hr.vn.si.d02.report'].create({
            'month': self.month,
            'year': self.year,
            'monthly_list_id': monthly_list.id,
        })
        # Tạo dòng chi tiết
        report.action_generate_lines()

        # Xuất file Excel
        file_data, filename = self._generate_excel(report)
        report.write({
            'export_file': file_data,
            'export_filename': filename,
        })

        return {
            'type': 'ir.actions.act_window',
            'name': _('Báo cáo D02-LT'),
            'res_model': 'hr.vn.si.d02.report',
            'res_id': report.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _generate_excel(self, report):
        if not openpyxl:
            raise UserError(
                _('Thư viện openpyxl chưa được cài đặt.')
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'D02-LT'

        # Header
        header_font = Font(bold=True, size=12)
        ws.merge_cells('A1:G1')
        ws['A1'] = f'DANH SÁCH LAO ĐỘNG THAM GIA BHXH, BHYT, BHTN'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Tháng {self.month}/{self.year}'
        ws['A2'].alignment = Alignment(horizontal='center')

        # Column headers
        headers = [
            'STT', 'Họ và tên', 'Số sổ BHXH',
            'Loại biến động', 'Mức lương cũ', 'Mức lương mới',
            'Từ ngày',
        ]
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        change_type_labels = {
            'increase': 'Tăng',
            'decrease': 'Giảm',
            'adjust': 'Điều chỉnh',
        }
        for idx, line in enumerate(report.line_ids, 1):
            row = idx + 4
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=line.full_name).border = thin_border
            ws.cell(row=row, column=3, value=line.bhxh_number or '').border = thin_border
            ws.cell(
                row=row, column=4,
                value=change_type_labels.get(line.change_type, ''),
            ).border = thin_border
            cell_old = ws.cell(row=row, column=5, value=line.old_salary)
            cell_old.border = thin_border
            cell_old.number_format = '#,##0'
            cell_new = ws.cell(row=row, column=6, value=line.new_salary)
            cell_new.border = thin_border
            cell_new.number_format = '#,##0'
            ws.cell(
                row=row, column=7,
                value=line.effective_date.strftime('%d/%m/%Y') if line.effective_date else '',
            ).border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 14

        # Save
        output = io.BytesIO()
        wb.save(output)
        file_data = base64.b64encode(output.getvalue())
        filename = f'D02_LT_T{self.month}_{self.year}.xlsx'
        return file_data, filename
