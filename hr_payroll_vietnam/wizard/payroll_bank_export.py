import base64
import io
import xml.etree.ElementTree as XmlET

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side
except ImportError:
    openpyxl = None


class HrVnBankExport(models.TransientModel):
    _name = 'hr.vn.bank.export'
    _description = 'Xuất file ngân hàng'

    payslip_run_id = fields.Many2one('hr.payslip.run', string='Đợt lương', required=True)
    bank_format = fields.Selection([
        ('vcb', 'Vietcombank (Excel)'), ('tcb', 'Techcombank (XML)'),
        ('mb', 'MB Bank (Excel)'), ('generic', 'Chung (CSV)'),
    ], string='Định dạng', required=True, default='vcb')
    export_file = fields.Binary(string='File xuất', readonly=True)
    export_filename = fields.Char(string='Tên file')

    def action_export(self):
        self.ensure_one()
        run = self.payslip_run_id
        if not run.slip_ids:
            raise UserError(_('Đợt lương chưa có phiếu lương.'))
        data = self._collect_data(run)
        method = {'vcb': self._export_vcb, 'tcb': self._export_tcb,
                  'mb': self._export_mb, 'generic': self._export_generic}
        file_data, filename = method[self.bank_format](data, run)
        self.write({'export_file': file_data, 'export_filename': filename})
        return {
            'type': 'ir.actions.act_window', 'name': _('Tải file'),
            'res_model': 'hr.vn.bank.export', 'res_id': self.id,
            'view_mode': 'form', 'target': 'new',
        }

    def _collect_data(self, run):
        data = []
        for slip in run.slip_ids:
            emp = slip.employee_id
            net_line = slip.line_ids.filtered(lambda l: l.code == 'NET')
            net = net_line[0].total if net_line else 0
            if net <= 0:
                continue
            data.append({
                'name': emp.name, 'bank_account': emp.bank_account_number or '',
                'bank_name': emp.bank_name or '', 'amount': net,
                'note': f'Lương T{run.date_start.month}/{run.date_start.year}',
            })
        return data

    def _make_excel(self, headers, data, run, prefix):
        if not openpyxl:
            raise UserError(_('Cần cài openpyxl.'))
        wb = openpyxl.Workbook()
        ws = wb.active
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
            c.border = thin
        for idx, row in enumerate(data, 1):
            r = idx + 1
            ws.cell(r, 1, idx).border = thin
            ws.cell(r, 2, row['name']).border = thin
            ws.cell(r, 3, row['bank_account']).border = thin
            ws.cell(r, 4, row['bank_name']).border = thin
            c = ws.cell(r, 5, row['amount'])
            c.border = thin
            c.number_format = '#,##0'
            ws.cell(r, 6, row['note']).border = thin
        output = io.BytesIO()
        wb.save(output)
        fn = f'{prefix}_T{run.date_start.month}_{run.date_start.year}.xlsx'
        return base64.b64encode(output.getvalue()), fn

    def _export_vcb(self, data, run):
        return self._make_excel(
            ['STT', 'Họ và tên', 'Số TK', 'Ngân hàng', 'Số tiền', 'Ghi chú'],
            data, run, 'VCB_Payment')

    def _export_mb(self, data, run):
        return self._make_excel(
            ['STT', 'Tên người nhận', 'Số TK', 'NH nhận', 'Số tiền', 'Nội dung'],
            data, run, 'MB_Payment')

    def _export_tcb(self, data, run):
        root = XmlET.Element('PaymentBatch')
        root.set('Bank', 'Techcombank')
        root.set('Date', str(run.date_end))
        for idx, row in enumerate(data, 1):
            p = XmlET.SubElement(root, 'Payment')
            XmlET.SubElement(p, 'No').text = str(idx)
            XmlET.SubElement(p, 'AccountName').text = row['name']
            XmlET.SubElement(p, 'AccountNumber').text = row['bank_account']
            XmlET.SubElement(p, 'BankName').text = row['bank_name']
            XmlET.SubElement(p, 'Amount').text = str(int(row['amount']))
            XmlET.SubElement(p, 'Description').text = row['note']
        output = io.BytesIO()
        tree = XmlET.ElementTree(root)
        XmlET.indent(tree, space='  ')
        tree.write(output, encoding='unicode', xml_declaration=True)
        fn = f'TCB_Payment_T{run.date_start.month}_{run.date_start.year}.xml'
        return base64.b64encode(output.getvalue().encode('utf-8')), fn

    def _export_generic(self, data, run):
        import csv
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        writer.writerow(['Họ và tên', 'Số tài khoản', 'Ngân hàng', 'Chi nhánh', 'Số tiền'])
        for row in data:
            writer.writerow([row['name'], row['bank_account'], row['bank_name'], '', int(row['amount'])])
        fn = f'Payment_T{run.date_start.month}_{run.date_start.year}.csv'
        return base64.b64encode(output.getvalue().encode('utf-8')), fn
