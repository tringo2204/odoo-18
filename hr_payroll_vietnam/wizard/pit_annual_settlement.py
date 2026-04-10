import base64
import io

from odoo import fields, models, _
from odoo.exceptions import UserError
from odoo.addons.hr_payroll_vietnam.models.vn_tax_engine import calculate_pit_progressive

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None


class PitAnnualSettlement(models.TransientModel):
    _name = 'hr.vn.pit.annual.settlement'
    _description = 'Quyết toán thuế TNCN năm (05/QTT-TNCN)'

    year = fields.Integer(
        string='Năm quyết toán', required=True,
        default=lambda self: fields.Date.today().year - 1,
    )
    company_id = fields.Many2one(
        'res.company', string='Công ty', required=True,
        default=lambda self: self.env.company,
    )
    employee_ids = fields.Many2many(
        'hr.employee', string='Nhân viên',
        help='Để trống = tất cả NV có phiếu lương trong năm',
    )
    export_file = fields.Binary(string='File xuất', readonly=True)
    export_filename = fields.Char()

    # Summary
    total_employees = fields.Integer(string='Số NV', readonly=True)
    total_gross = fields.Float(string='Tổng thu nhập', readonly=True)
    total_pit_paid = fields.Float(string='Tổng thuế đã nộp', readonly=True)
    total_pit_annual = fields.Float(string='Thuế phải nộp cả năm', readonly=True)
    total_difference = fields.Float(string='Chênh lệch', readonly=True)

    def action_compute(self):
        self.ensure_one()
        results = self._compute_annual_pit()
        self.write({
            'total_employees': len(results),
            'total_gross': sum(r['total_gross'] for r in results),
            'total_pit_paid': sum(r['pit_paid'] for r in results),
            'total_pit_annual': sum(r['pit_annual'] for r in results),
            'total_difference': sum(r['difference'] for r in results),
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_export(self):
        self.ensure_one()
        results = self._compute_annual_pit()
        if not results:
            raise UserError(_('Không có dữ liệu để xuất.'))
        file_data, filename = self._generate_excel(results)
        self.write({
            'export_file': file_data,
            'export_filename': filename,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _compute_annual_pit(self):
        """Tính quyết toán thuế TNCN năm theo biểu năm."""
        Payslip = self.env['hr.payslip']
        domain = [
            ('date_from', '>=', f'{self.year}-01-01'),
            ('date_to', '<=', f'{self.year}-12-31'),
            ('state', '=', 'done'),
            ('company_id', '=', self.company_id.id),
        ]
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))

        payslips = Payslip.search(domain)
        if not payslips:
            return []

        # Load PIT config
        deduction = self.env['hr.vn.personal.deduction'].search(
            [('year', '=', self.year)], limit=1,
        )
        self_ded = deduction.self_deduction if deduction else 11000000
        dep_ded = deduction.dependent_deduction if deduction else 4400000

        brackets = self.env['hr.vn.pit.bracket'].search(
            [('year', '=', self.year)], order='bracket_no',
        )
        bracket_list = [(b.income_from, b.income_to, b.tax_rate) for b in brackets]
        if not bracket_list:
            bracket_list = [
                (0, 5000000, 5), (5000000, 10000000, 10),
                (10000000, 18000000, 15), (18000000, 32000000, 20),
                (32000000, 52000000, 25), (52000000, 80000000, 30),
                (80000000, 0, 35),
            ]

        # Group by employee
        emp_data = {}
        for slip in payslips:
            emp_id = slip.employee_id.id
            if emp_id not in emp_data:
                emp_data[emp_id] = {
                    'employee': slip.employee_id,
                    'total_gross': 0, 'total_insurance': 0,
                    'pit_paid': 0, 'months': 0,
                }
            lines = {l.code: l.total for l in slip.line_ids}
            emp_data[emp_id]['total_gross'] += lines.get('GROSS', 0)
            emp_data[emp_id]['total_insurance'] += abs(lines.get('BHXH_EE', 0)) + abs(lines.get('BHYT_EE', 0)) + abs(lines.get('BHTN_EE', 0))
            emp_data[emp_id]['pit_paid'] += abs(lines.get('PIT', 0))
            emp_data[emp_id]['months'] += 1

        # Calculate annual PIT per employee
        results = []
        for emp_id, data in emp_data.items():
            emp = data['employee']
            dep_count = self.env['hr.vn.dependent'].search_count([
                ('employee_id', '=', emp_id),
                ('status', '=', 'approved'),
            ])
            total_deduction = (self_ded * data['months']) + (dep_ded * dep_count * data['months'])
            taxable = data['total_gross'] - data['total_insurance'] - total_deduction
            if taxable < 0:
                taxable = 0

            # Annual PIT: taxable / 12 → monthly bracket → × 12
            monthly_taxable = taxable / 12 if taxable else 0
            pit_monthly = calculate_pit_progressive(monthly_taxable, bracket_list)
            pit_annual = pit_monthly * 12

            results.append({
                'employee': emp,
                'tax_id': emp.tax_id or '',
                'months': data['months'],
                'total_gross': data['total_gross'],
                'total_insurance': data['total_insurance'],
                'total_deduction': total_deduction,
                'taxable': taxable,
                'pit_paid': data['pit_paid'],
                'pit_annual': pit_annual,
                'difference': data['pit_paid'] - pit_annual,
            })
        return results

    def _generate_excel(self, results):
        if not openpyxl:
            raise UserError(_('Cần cài openpyxl.'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '05_QTT_TNCN'
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        ws.merge_cells('A1:K1')
        ws['A1'] = f'QUYẾT TOÁN THUẾ THU NHẬP CÁ NHÂN NĂM {self.year}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['STT', 'Họ và tên', 'MST', 'Số tháng', 'Tổng thu nhập',
                   'Tổng BH', 'Tổng giảm trừ', 'TNCT', 'Thuế đã nộp',
                   'Thuế phải nộp', 'Chênh lệch']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        for idx, r in enumerate(results, 1):
            row = idx + 3
            ws.cell(row, 1, idx).border = thin
            ws.cell(row, 2, r['employee'].name).border = thin
            ws.cell(row, 3, r['tax_id']).border = thin
            ws.cell(row, 4, r['months']).border = thin
            for col, key in enumerate(['total_gross', 'total_insurance', 'total_deduction',
                                        'taxable', 'pit_paid', 'pit_annual', 'difference'], 5):
                c = ws.cell(row, col, r[key])
                c.border = thin
                c.number_format = '#,##0'

        for col in 'ABCDEFGHIJK':
            ws.column_dimensions[col].width = 18

        output = io.BytesIO()
        wb.save(output)
        filename = f'05_QTT_TNCN_{self.year}.xlsx'
        return base64.b64encode(output.getvalue()), filename
